import openpyxl
import sqlite3
from datetime import datetime

# Excel dosyasını oku
wb = openpyxl.load_workbook('c:/satinalma/procurement_system_nodejs/data/data4.xlsx')
ws = wb.active

# Veritabanı bağlantısı
conn = sqlite3.connect('c:/satinalma/procurement_system_nodejs/database.sqlite')
cursor = conn.cursor()

# Sütun başlıkları
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

# Eşleme
header_map = {
    'Müşteri/tedarikçi ismi': 'name',
    'E-posta': 'email',
    'Adres': 'address',
    'Telefon': 'phone',
    'Vergi dairesi': 'tax_office',
    'Vergi numarası/TC kimlik no': 'tax_number',
    # Diğer başlıklar eklenebilir
}

# Varsayılanlar
default_fields = {
    'code': None,  # Otomatik üretilecek
    'business_type': 'both',
    'status': 'active',
    'currency': 'TRY',
    'timezone': 'Europe/Istanbul',
    'type': 'customer',
}

# Satırları işle
for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
    data = {k: None for k in header_map.values()}
    for excel_col, db_col in header_map.items():
        if excel_col in headers:
            data[db_col] = row[headers.index(excel_col)]
    # code alanı otomatik üret
    data['code'] = f"COMP{idx:03d}"
    # Diğer varsayılanları ekle
    for k, v in default_fields.items():
        if k not in data or data[k] is None:
            data[k] = v
    # created_at ve updated_at ekle
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    # Tekrarlayan kayıt kontrolü (name veya code)
    cursor.execute("SELECT COUNT(*) FROM companies WHERE name = ? OR code = ?", (data['name'], data['code']))
    exists = cursor.fetchone()[0]
    if exists:
        print(f"Atlandı (tekrar): {data['name']} / {data['code']}")
        continue
    # SQL insert
    cursor.execute('''
        INSERT INTO companies (name, code, address, email, phone, tax_office, tax_number, business_type, status, currency, timezone, type, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data['name'], data['code'], data['address'], data['email'], data['phone'],
        data['tax_office'], data['tax_number'], data['business_type'], data['status'],
        data['currency'], data['timezone'], data['type'], now, now
    ))

conn.commit()
conn.close()
print('Aktarım tamamlandı.')
