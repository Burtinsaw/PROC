import openpyxl

wb = openpyxl.load_workbook('c:/satinalma/procurement_system_nodejs/data/data1.xlsx')
ws = wb.active

# Sütun başlıklarını al
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print('Başlıklar:', headers)

# İlk 5 satırı yazdır
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=1):
    print(f"Satır {i}: {row}")
